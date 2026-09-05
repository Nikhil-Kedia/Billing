"""
inventory_io.py - exporting the item list to a spreadsheet, and reading
one back in.

WHY

Four hundred items were originally loaded by running a throwaway script
(import_data.py) from a terminal against a hard-coded filename. There was
no way to get the list back OUT at all. A shopkeeper who wants to put
prices up 5% across a category, or hand the catalogue to an accountant,
had no route that did not involve a developer.

WHAT IT DOES

  Export  - every item to .csv or .xlsx, in exactly the columns the
            importer accepts, so an export can be edited in Excel and
            imported straight back.
  Import  - adds items that are new, and asks skip/overwrite for the ones
            that already exist.

WHAT IT DELIBERATELY REUSES

The skip/overwrite behaviour is NOT written again here. backup_restore
already has a conflict model (Conflict, with a `.resolution` the UI
fills in) and gui/backup_dialogs has the review dialog with its
"skip all" / "overwrite all" buttons. This module produces the same
shapes and hands them to that dialog, so imported spreadsheets and
imported backups behave identically and there is one place to fix.

Every incoming row goes through validation.py, exactly as typed input
does - a spreadsheet is untrusted input like any other file.
"""

import csv
import io
import os

import backup_restore as br
import database as db
import validation

# The importer accepts these, and the exporter writes them, in this
# order. Keeping the two in one list is what makes a round trip work.
COLUMNS = [
    ("item_code", "Item Code"),
    ("name", "Product Name"),
    ("category", "Category"),
    ("unit", "Unit"),
    ("price", "Price"),
    ("quantity", "Stock Qty"),
    ("low_stock_threshold", "Low Stock Alert At"),
    ("pack_size", "Pack Size"),
    ("pack_unit_name", "Pack Unit Name"),
]

FIELDS = [key for key, _label in COLUMNS]
HEADERS = [label for _key, label in COLUMNS]

# Accepted spellings for each column, so a file that came from somewhere
# else (or from the old import_data.py, which used bare lowercase names)
# still lands in the right place.
ALIASES = {
    "item_code": {"item_code", "item code", "code", "sku", "item"},
    "name": {"name", "product name", "product", "item name", "description"},
    "category": {"category", "group", "type"},
    "unit": {"unit", "uom", "units"},
    "price": {"price", "rate", "mrp", "selling price", "price/unit"},
    "quantity": {"quantity", "qty", "stock", "stock qty", "current stock"},
    "low_stock_threshold": {"low_stock_threshold", "low stock alert at",
                            "low stock", "reorder level", "alert at"},
    "pack_size": {"pack_size", "pack size", "carton size"},
    "pack_unit_name": {"pack_unit_name", "pack unit name", "pack unit", "carton"},
}

MAX_IMPORT_ROWS = 50_000


class ImportError_(Exception):
    """Message is safe to show the user."""


# =====================================================================
# EXPORT
# =====================================================================

def export_items(path):
    """Writes every item to `path`. Format follows the extension."""
    items = db.get_all_items()
    rows = [[_export_value(item.get(field)) for field in FIELDS] for item in items]

    if path.lower().endswith(".xlsx"):
        _write_xlsx(path, rows)
    else:
        _write_csv(path, rows)
    return len(rows)


def _export_value(value):
    if value is None:
        return ""
    return value


def _write_csv(path, rows):
    # utf-8-sig: Excel opens a plain UTF-8 CSV as mojibake unless there is
    # a BOM, and shop catalogues contain non-ASCII names.
    with io.open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(rows)


def _write_xlsx(path, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    book = Workbook()
    sheet = book.active
    sheet.title = "Inventory"
    sheet.append(HEADERS)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"          # headers stay put while scrolling

    for row in rows:
        sheet.append(row)

    for index, header in enumerate(HEADERS, start=1):
        longest = max([len(str(header))] +
                      [len(str(r[index - 1])) for r in rows[:500]] or [0])
        sheet.column_dimensions[get_column_letter(index)].width = min(longest + 4, 40)

    book.save(path)


# =====================================================================
# IMPORT
# =====================================================================

def read_rows(path):
    """Reads a spreadsheet into a list of cleaned dicts.

    Raises ImportError_ with a readable message for anything a user can
    actually fix - a missing name column, an unreadable file, a sheet
    with nothing in it.
    """
    if not os.path.isfile(path):
        raise ImportError_("That file could not be found.")

    if path.lower().endswith((".xlsx", ".xlsm")):
        table = _read_xlsx(path)
    else:
        table = _read_csv(path)

    if not table:
        raise ImportError_("That file is empty.")

    headers = [str(h or "").strip().lower() for h in table[0]]
    mapping = {}
    for field, names in ALIASES.items():
        for index, header in enumerate(headers):
            if header in names:
                mapping[field] = index
                break

    if "name" not in mapping:
        raise ImportError_(
            "Could not find a product-name column." + chr(10) + chr(10)
            + "The first row must be headings, and one of them must be "
              "'Product Name' (or 'name'). Export your inventory first to "
              "see the exact format.")

    rows = []
    for line_no, raw in enumerate(table[1:], start=2):
        if len(rows) >= MAX_IMPORT_ROWS:
            break
        if not any(str(cell or "").strip() for cell in raw):
            continue                                  # blank spreadsheet row
        try:
            rows.append(_clean_row(raw, mapping, line_no))
        except validation.ValidationError as e:
            raise ImportError_(f"Row {line_no}: {e}")
    if not rows:
        raise ImportError_("That file has headings but no item rows.")
    return rows


def _cell(raw, mapping, field):
    index = mapping.get(field)
    if index is None or index >= len(raw):
        return ""
    value = raw[index]
    return "" if value is None else str(value).strip()


def _clean_row(raw, mapping, line_no):
    """Runs a spreadsheet row through the same validation as typed input."""
    name = validation.item_name(_cell(raw, mapping, "name"))
    return {
        "item_code": validation.item_code(_cell(raw, mapping, "item_code")),
        "name": name,
        "category": validation.category(_cell(raw, mapping, "category")),
        "unit": validation.unit(_cell(raw, mapping, "unit")) or "pcs",
        "price": validation.money(_cell(raw, mapping, "price") or 0,
                                  f"Price for '{name}'"),
        "quantity": validation.quantity(_cell(raw, mapping, "quantity") or 0,
                                        f"Stock for '{name}'", allow_zero=True),
        "low_stock_threshold": validation.quantity(
            _cell(raw, mapping, "low_stock_threshold") or 0,
            f"Low-stock level for '{name}'", allow_zero=True),
        "pack_size": validation.pack_size(_cell(raw, mapping, "pack_size")),
        "pack_unit_name": validation.unit(_cell(raw, mapping, "pack_unit_name")),
        "_line": line_no,
    }


def _read_csv(path):
    with io.open(path, "r", encoding="utf-8-sig", errors="replace") as handle:
        content = handle.read()
    if not content.strip():
        return []
    # Sniff the delimiter rather than assuming a comma: the shop's own
    # existing file is semicolon-separated, which is what Excel writes in
    # locales where the comma is the decimal separator.
    first_line = content.splitlines()[0]
    delimiter = ","
    for candidate in (";", "\t", "|"):
        if first_line.count(candidate) > first_line.count(delimiter):
            delimiter = candidate
    return [row for row in csv.reader(io.StringIO(content), delimiter=delimiter)]


def _read_xlsx(path):
    from openpyxl import load_workbook
    # read_only + data_only: we want values, not formulas, and not the
    # whole workbook in memory.
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        book.close()


def scan(rows):
    """Splits incoming rows into new ones and conflicts.

    Matches on item_code first and falls back to name, which is UNIQUE in
    the schema - so an import cannot create a second item with a name
    that already exists, which would fail at the database anyway.

    Returns (plan, conflicts) shaped for gui.backup_dialogs.ConflictReviewDialog.
    """
    snapshot = db.items_snapshot()
    by_code = snapshot["by_code"]
    by_name = snapshot["by_name"]

    plan = []
    conflicts = []
    seen = set()

    for row in rows:
        existing = None
        if row["item_code"]:
            existing = by_code.get(row["item_code"])
        if existing is None:
            existing = by_name.get(row["name"])

        # A file that lists the same item twice would otherwise insert it
        # once and then fail on the UNIQUE constraint.
        key = (row["item_code"] or "", row["name"])
        if key in seen:
            continue
        seen.add(key)

        if existing is None:
            plan.append({"row": row, "conflict": None})
            continue

        # The description names the EXISTING item, not the incoming one.
        # Matching is by item_code first, so a spreadsheet row labelled
        # "KURKURE" carrying code 1 matches whatever item already holds
        # code 1 - which may be something else entirely. Showing only the
        # incoming name would let someone overwrite the wrong item without
        # ever seeing which one they were replacing.
        matched_on = "code " + row["item_code"] if (
            row["item_code"] and by_code.get(row["item_code"]) is existing) else "name"
        if existing["name"] != row["name"]:
            headline = f"{existing['name']}  ->  rename to  {row['name']}"
        else:
            headline = existing["name"]
        conflict = br.Conflict(
            "items", f"items:{row['name']}",
            f"{headline}   (matched on {matched_on})   "
            f"price {theme_currency(existing['price'])} -> {theme_currency(row['price'])}",
            dict(existing), row, existing["id"])
        conflicts.append(conflict)
        plan.append({"row": row, "conflict": conflict})

    return plan, conflicts


def theme_currency(value):
    """Local formatter so this module does not import the GUI theme."""
    try:
        return f"Rs. {float(value):,.2f}"
    except (TypeError, ValueError):
        return "Rs. 0.00"


def apply(plan):
    """Applies a scanned plan once every conflict has a resolution.

    Returns {"added": n, "updated": n, "skipped": n}.
    """
    counts = {"added": 0, "updated": 0, "skipped": 0}

    for entry in plan:
        row, conflict = entry["row"], entry["conflict"]
        if conflict is None:
            db.add_item(
                item_code=row["item_code"], name=row["name"],
                category=row["category"], unit=row["unit"], price=row["price"],
                quantity=row["quantity"],
                low_stock_threshold=row["low_stock_threshold"],
                pack_size=row["pack_size"], pack_unit_name=row["pack_unit_name"])
            counts["added"] += 1
        elif conflict.resolution == "overwrite":
            existing = dict(conflict.existing_row)
            db.update_item(
                item_id=conflict.existing_id,
                item_code=row["item_code"] or existing.get("item_code"),
                name=row["name"], category=row["category"], unit=row["unit"],
                price=row["price"],
                # Stock is NOT taken from the spreadsheet on an overwrite.
                # A price list edited in Excel is almost never an accurate
                # stock count, and silently overwriting live quantities
                # with a stale column is not recoverable.
                quantity=existing.get("quantity", 0),
                low_stock_threshold=row["low_stock_threshold"],
                pack_size=row["pack_size"], pack_unit_name=row["pack_unit_name"])
            counts["updated"] += 1
        else:
            counts["skipped"] += 1

    db.bump("items")
    return counts
